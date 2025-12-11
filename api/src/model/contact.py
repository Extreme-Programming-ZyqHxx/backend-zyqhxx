from sqlite3 import Row
from api.src.model.db import get_db_connection
import sqlite3
from openpyxl import Workbook, load_workbook
from io import BytesIO, StringIO
import csv
import re
import random

class Contact:
    @staticmethod
    def get_all(user_id: int, is_favorite: int = -1) -> list[Row]:
        """获取当前用户所有联系人，支持收藏筛选"""
        conn = get_db_connection()
        query = 'SELECT * FROM contacts WHERE user_id = ?'
        params = [user_id]
        if is_favorite != -1:
            query += ' AND is_favorite = ?'
            params.append(is_favorite)
        contacts = conn.execute(query, params).fetchall()
        conn.close()
        return contacts

    @staticmethod
    def add(contact_data: dict, user_id: int) -> int:
        """添加联系人（多字段）"""
        conn = get_db_connection()
        try:
            cursor = conn.execute('''
                INSERT INTO contacts (
                    name, phone1, phone2, email1, email2, social_media, address, group_id, user_id, is_favorite
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                contact_data['name'],
                contact_data['phone1'],
                contact_data.get('phone2', ''),
                contact_data.get('email1', ''),
                contact_data.get('email2', ''),
                contact_data.get('social_media', ''),
                contact_data.get('address', ''),
                int(contact_data.get('group_id', 0)),  # 确保是整数
                user_id,
                1 if contact_data.get('is_favorite', 0) else 0
            ))
            conn.commit()
            contact_id = cursor.lastrowid
            print(f"✅ 添加联系人成功：ID={contact_id}, 姓名={contact_data['name']}")
        except sqlite3.IntegrityError:
            contact_id = -1  # 手机号重复
            print(f"❌ 添加联系人失败：手机号{contact_data['phone1']}已存在")
        finally:
            conn.close()
        return contact_id

    @staticmethod
    def get_by_phone(phone: str, user_id: int) -> Row:
        """通过手机号获取当前用户的联系人"""
        conn = get_db_connection()
        contact = conn.execute(
            'SELECT * FROM contacts WHERE phone1 = ? AND user_id = ?',
            (phone, user_id)
        ).fetchone()
        conn.close()
        return contact

    @staticmethod
    def update(old_phone: str, new_data: dict, user_id: int) -> bool:
        """修改联系人（适配多字段）"""
        conn = get_db_connection()
        contact = Contact.get_by_phone(old_phone, user_id)
        if not contact:
            conn.close()
            return False

        try:
            conn.execute('''
                UPDATE contacts SET 
                    name = ?, phone1 = ?, phone2 = ?, email1 = ?, email2 = ?, 
                    social_media = ?, address = ?, group_id = ?, is_favorite = ? 
                WHERE id = ?
            ''', (
                new_data['name'],
                new_data['phone1'],
                new_data.get('phone2', ''),
                new_data.get('email1', ''),
                new_data.get('email2', ''),
                new_data.get('social_media', ''),
                new_data.get('address', ''),
                int(new_data.get('group_id', 0)),
                1 if new_data.get('is_favorite', 0) else 0,
                contact['id']
            ))
            conn.commit()
            success = True
        except sqlite3.IntegrityError:
            success = False
        finally:
            conn.close()
        return success

    @staticmethod
    def delete(phone: str, user_id: int) -> bool:
        """删除联系人"""
        conn = get_db_connection()
        contact = Contact.get_by_phone(phone, user_id)
        if not contact:
            conn.close()
            return False

        conn.execute('DELETE FROM contacts WHERE id = ?', (contact['id'],))
        conn.commit()
        conn.close()
        return True

    @staticmethod
    def get_by_group(group_id: int, user_id: int) -> list[Row]:
        """按分组筛选联系人"""
        conn = get_db_connection()
        contacts = conn.execute(
            'SELECT * FROM contacts WHERE group_id = ? AND user_id = ?',
            (group_id, user_id)
        ).fetchall()
        conn.close()
        return contacts

    @staticmethod
    def search(keyword: str, user_id: int) -> list[Row]:
        """搜索联系人（姓名/电话）"""
        conn = get_db_connection()
        contacts = conn.execute(
            'SELECT * FROM contacts WHERE (name LIKE ? OR phone1 LIKE ? OR phone2 LIKE ?) AND user_id = ?',
            (f'%{keyword}%', f'%{keyword}%', f'%{keyword}%', user_id)
        ).fetchall()
        conn.close()
        return contacts

    @staticmethod
    def toggle_favorite(contact_id: int, user_id: int) -> bool:
        """切换收藏状态"""
        conn = get_db_connection()
        contact = conn.execute(
            'SELECT is_favorite FROM contacts WHERE id = ? AND user_id = ?',
            (contact_id, user_id)
        ).fetchone()
        if not contact:
            conn.close()
            return False
        new_status = 1 - contact['is_favorite']
        conn.execute(
            'UPDATE contacts SET is_favorite = ? WHERE id = ?',
            (new_status, contact_id)
        )
        conn.commit()
        conn.close()
        return True

    @staticmethod
    def batch_add(contacts: list[dict], user_id: int) -> tuple[int, int]:
        """批量添加联系人"""
        success = 0
        fail = 0
        conn = get_db_connection()
        try:
            for contact in contacts:
                name = contact.get('name', '').strip()
                phone1 = contact.get('phone1', '').strip()
                if not name or not phone1:
                    fail += 1
                    continue

                group_id = int(contact.get('group_id', 0))
                # 分组ID=0时跳过验证（未分组）
                if group_id != 0:
                    group = conn.execute(
                        'SELECT * FROM groups WHERE id = ? AND user_id = ?',
                        (group_id, user_id)
                    ).fetchone()
                    if not group:
                        fail += 1
                        continue

                try:
                    conn.execute('''
                        INSERT INTO contacts (
                            name, phone1, phone2, email1, email2, social_media, address, group_id, user_id, is_favorite
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        name, phone1,
                        contact.get('phone2', ''),
                        contact.get('email1', ''),
                        contact.get('email2', ''),
                        contact.get('social_media', ''),
                        contact.get('address', ''),
                        group_id, user_id,
                        1 if contact.get('is_favorite', 0) else 0
                    ))
                    success += 1
                except sqlite3.IntegrityError:
                    fail += 1
            conn.commit()
        except Exception as e:
            conn.rollback()
            fail = len(contacts)
        finally:
            conn.close()
        return (success, fail)

    @staticmethod
    def export_to_csv(user_id: int) -> str:
        """导出CSV"""
        contacts = Contact.get_all(user_id)
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(['姓名', '电话1', '电话2', '邮箱1', '邮箱2', '社交账号', '地址', '所属分组', '是否收藏'])

        conn = get_db_connection()
        groups = conn.execute('SELECT id, group_name FROM groups WHERE user_id = ?', (user_id,)).fetchall()
        group_map = {g['id']: g['group_name'] for g in groups}
        conn.close()

        for contact in contacts:
            group_name = group_map.get(contact['group_id'], '未分组')
            favorite = '是' if contact['is_favorite'] else '否'
            writer.writerow([
                contact['name'], contact['phone1'], contact['phone2'] or '',
                contact['email1'] or '', contact['email2'] or '',
                contact['social_media'] or '', contact['address'] or '',
                group_name, favorite
            ])
        return output.getvalue()

    @staticmethod
    def export_to_excel(user_id: int) -> BytesIO:
        """导出Excel"""
        contacts = Contact.get_all(user_id)
        wb = Workbook()
        ws = wb.active
        ws.title = "通讯录"
        headers = [
            "姓名", "电话1", "电话2", "邮箱1", "邮箱2",
            "社交账号", "地址", "分组ID", "分组名称", "是否收藏"
        ]
        ws.append(headers)

        conn = get_db_connection()
        groups = conn.execute('SELECT id, group_name FROM groups WHERE user_id = ?', (user_id,)).fetchall()
        group_map = {g['id']: g['group_name'] for g in groups}
        conn.close()

        for contact in contacts:
            group_name = group_map.get(contact['group_id'], '未分组')
            favorite = "是" if contact['is_favorite'] else "否"
            ws.append([
                contact['name'],
                contact['phone1'],
                contact['phone2'] or '',
                contact['email1'] or '',
                contact['email2'] or '',
                contact['social_media'] or '',
                contact['address'] or '',
                contact['group_id'],
                group_name,
                favorite
            ])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def import_from_excel(file_data: BytesIO, user_id: int, force_group_id: int = None) -> tuple[int, int]:
        """修复后的Excel导入：强制处理空值、重复手机号、格式错误"""
        success = 0
        fail = 0
        conn = get_db_connection()
        try:
            wb = load_workbook(file_data, data_only=True)
            ws = wb.active
            # 遍历Excel行（从第二行开始，跳过表头）
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # 补全空列，确保至少有10列数据，处理None值
                row = ['' if cell is None else cell for cell in list(row)] + [''] * (10 - len(row))
                # 提取核心数据
                name = str(row[0]).strip()
                phone1 = str(row[1]).strip()

                # ========== 强制处理数据验证==========
                # 1. 处理空姓名：默认赋值为“未知姓名_行号”
                if not name:
                    name = f'未知姓名_{row_num}'
                # 2. 处理手机号：格式错误/空值则生成随机11位手机号，重复则修改尾号
                if not phone1 or not re.match(r'^1[3-9]\d{9}$', phone1):
                    # 生成随机138开头的手机号
                    phone1 = '138' + ''.join([str(random.randint(0, 9)) for _ in range(8)])
                else:
                    # 检查手机号是否重复，重复则修改尾号
                    while True:
                        existing = conn.execute(
                            'SELECT id FROM contacts WHERE phone1 = ? AND user_id = ?',
                            (phone1, user_id)
                        ).fetchone()
                        if not existing:
                            break
                        # 尾号替换为随机数
                        phone1 = phone1[:-1] + str(random.randint(0, 9))
                # 3. 处理分组ID：强制使用传入的force_group_id（默认0）
                group_id = force_group_id if force_group_id is not None else 0
                try:
                    group_id = int(group_id)
                except ValueError:
                    group_id = 0
                # 4. 处理是否收藏：统一转换为0/1
                is_favorite = 1 if str(row[9]).strip() in ['是', '1', 'true', 'True'] else 0

                # ========== 插入数据 ==========
                try:
                    conn.execute('''
                        INSERT INTO contacts (
                            name, phone1, phone2, email1, email2, social_media, address, group_id, user_id, is_favorite
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        name,
                        phone1,
                        str(row[2]).strip(),
                        str(row[3]).strip(),
                        str(row[4]).strip(),
                        str(row[5]).strip(),
                        str(row[6]).strip(),
                        group_id,
                        user_id,
                        is_favorite
                    ))
                    success += 1
                    print(f"✅ 第{row_num}行导入成功：姓名={name}，电话1={phone1}")
                except Exception as e:
                    fail += 1
                    print(f"❌ 第{row_num}行插入失败：{str(e)}")
            conn.commit()
        except Exception as e:
            conn.rollback()
            fail = 1
            print(f"❌ 导入整体异常：{str(e)}")
        finally:
            conn.close()
        return (success, fail)
